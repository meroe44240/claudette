#!/usr/bin/env python3
"""Build the consolidated market-mapping XLSX from the day's CSV deliverables."""
import csv
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
GREEN = PatternFill("solid", fgColor="C6EFCE")
RED = PatternFill("solid", fgColor="FFC7CE")

SHEETS = [
    ("Valentin - Finance", "finance"),
    ("Valentin - Hospitality", "hospitality"),
    ("Alexis - Industrie", "industrie"),
    ("Louis - Sales SaaS", "sales_saas"),
    ("Meroe - Sales", "sales"),
]


def read_csv(path):
    if not os.path.exists(path):
        return None
    with open(path, encoding="utf-8-sig", newline="") as fh:
        return list(csv.reader(fh, delimiter=";"))


def style_sheet(ws, rows):
    if not rows:
        return
    tags_idx = rows[0].index("tags") if "tags" in rows[0] else None
    for row in rows:
        ws.append(row)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    if tags_idx is not None:
        col = tags_idx + 1
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=col)
            value = cell.value or ""
            cell.fill = RED if "CONTACT_NON_SOURCE" in value else GREEN
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        width = max(
            (len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[letter].width = min(max(width + 2, 10), 60)


def main(date):
    out_dir = "outputs"
    wb = Workbook()

    synth = wb.active
    synth.title = "Synthese"

    stats = []
    sheet_data = {}
    for title, key in SHEETS:
        rows = read_csv(os.path.join(out_dir, f"{key}_{date}.csv"))
        sheet_data[key] = rows
        if rows is None:
            stats.append((title, "FICHIER MANQUANT", "", "", ""))
            continue
        body = rows[1:]
        header = rows[0]
        ent = header.index("entreprise")
        tags = header.index("tags")
        uniq = len({r[ent] for r in body if len(r) > ent})
        sourced = sum(1 for r in body if len(r) > tags and "CONTACT_NON_SOURCE" not in r[tags])
        pct = f"{round(100 * sourced / len(body))}%" if body else "0%"
        stats.append((title, len(body), uniq, sourced, pct))

    synth.append(["Humanup.io - Market Mapping", date])
    synth.append([])
    synth.append(["Vertical", "Lignes", "Entreprises uniques", "Contacts sources", "% source"])
    for row in stats:
        synth.append(list(row))
    synth.append([])

    report = os.path.join(out_dir, f"rapport_synthese_{date}.md")
    if os.path.exists(report):
        synth.append(["Rapport de synthese"])
        with open(report, encoding="utf-8") as fh:
            for line in fh:
                synth.append([line.rstrip("\n")])

    for cell in synth[3]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    synth["A1"].font = Font(bold=True, size=14)
    synth.column_dimensions["A"].width = 60
    for letter in "BCDE":
        synth.column_dimensions[letter].width = 22

    for title, key in SHEETS:
        ws = wb.create_sheet(title[:31])
        rows = sheet_data.get(key)
        if rows:
            style_sheet(ws, rows)
        else:
            ws.append(["Fichier manquant pour cette verticale"])

    path = os.path.join(out_dir, f"humanup_market_mapping_{date}.xlsx")
    wb.save(path)
    print(f"XLSX ecrit: {path} ({len(wb.sheetnames)} onglets)")
    for row in stats:
        print("  ", row)


if __name__ == "__main__":
    main(sys.argv[1])
